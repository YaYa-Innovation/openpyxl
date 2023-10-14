from openpyxl import load_workbook
import datetime

# Load the Excel file
workbook = load_workbook('data.xlsx')

# Select the active worksheet
worksheet = workbook.active

# Get the maximum row count
max_row = worksheet.max_row

# Print the last 5 rows
for row in worksheet.iter_rows(min_row=max_row-4, max_row=max_row, values_only=True):
    date_obj = row[0]  # Extract the datetime object
    formatted_date = date_obj.strftime('%Y-%m-%d')  # Format date as YYYY-MM-DD
    values = row[1:]  # Extract values excluding the datetime
    
    print(formatted_date, values)

# Close the workbook
workbook.close()
