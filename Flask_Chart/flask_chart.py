from flask import Flask, render_template, jsonify, request, send_file
from openpyxl import *
import io


from datetime import datetime

app = Flask(__name__)

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/view')
def view():
    return render_template('view.html')

@app.route('/get_data')
def get_data():
    workbook = load_workbook('static/data.xlsx')
    sheet = workbook.active

    dates = []
    values1 = []
    values2 = []
    values3 = []
    values4 = []
    values5 = []
    values6 = []
    values7 = []
    values8 = []
    values9 = []
    values10 = []
    values11= []
    values12 = []
    values13= []
    values14= []
    values15= []
    values16 = []
    values17 = []

    for row in sheet.iter_rows(min_row=sheet.max_row-30, max_row=sheet.max_row, min_col=1, max_col=18):
        date = row[0].value
        dates.append(date.strftime('%d-%m-%Y'))
        values1.append(row[1].value)
        values2.append(row[2].value)
        values3.append(row[3].value)
        values4.append(row[4].value)
        values5.append(row[5].value)
        values6.append(row[6].value)
        values7.append(row[7].value)
        values8.append(row[8].value)
        values9.append(row[9].value)
        values10.append(row[10].value)
        values11.append(row[11].value)
        values12.append(row[12].value)
        values13.append(row[13].value)
        values14.append(row[14].value)
        values15.append(row[15].value)
        values16.append(row[16].value)
        values17.append(row[17].value)

    return jsonify(dates=dates, values1=values1, values2=values2, values3=values3,values4=values4,values5=values5,values6=values6,values7=values7,values8=values8,values9=values9,values10=values10,values11=values11,values12=values12,values13=values13,values14=values14,values15=values15,values16=values16,values17=values17)

@app.route('/get_filtered_data')
def get_filtered_data():
    workbook = load_workbook('static/data.xlsx')
    sheet = workbook.active

    start_date = request.args.get('startDate')
    end_date = request.args.get('endDate')

    dates = []
    values1 = []
    values2 = []
    values3 = []
    values4 = []
    values5 = []
    values6 = []
    values7 = []
    values8 = []
    values9 = []
    values10 = []
    values11= []
    values12 = []
    values13= []
    values14= []
    values15= []
    values16 = []
    values17 = []

#    return jsonify(dates=dates, values1=values1, values2=values2, values3=values3, values4=values4)
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=18):
        date = row[0].value
        date_str = date.strftime('%Y-%m-%d')  # Convert datetime to string

        if not start_date or not end_date or (not start_date and not end_date) or (start_date <= date_str <= end_date):
            dates.append(date_str)
            values1.append(row[1].value)
            values2.append(row[2].value)
            values3.append(row[3].value)
            values4.append(row[4].value)
            values5.append(row[5].value)
            values6.append(row[6].value)
            values7.append(row[7].value)
            values8.append(row[8].value)
            values9.append(row[9].value)
            values10.append(row[10].value)
            values11.append(row[11].value)
            values12.append(row[12].value)
            values13.append(row[13].value)
            values14.append(row[14].value)
            values15.append(row[15].value)
            values16.append(row[16].value)
            values17.append(row[17].value)

    return jsonify(dates=dates, values1=values1, values2=values2, values3=values3,values4=values4,values5=values5,values6=values6,values7=values7,values8=values8,values9=values9,values10=values10,values11=values11,values12=values12,values13=values13,values14=values14,values15=values15,values16=values16,values17=values17)

#    return jsonify(dates=dates, values1=values1, values2=values2, values3=values3, values4=values4)
@app.route('/export_excel')
def export_excel():
    workbook_out = Workbook()
    sheet_out = workbook_out.active

    # Write headers
    headers = ['Date', 'Value 1', 'Value 2', 'Value 3','values 4','values 5','values 6','values 7','values 8','values 9','values 10','Value 11', 'Value 12', 'Value 13','values 14','values 15','values 16','values 17']
    sheet_out.append(headers)

    # Retrieve data from the original file
    workbook = load_workbook('static/data.xlsx')
    sheet = workbook.active

    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=18):
        data_row = [cell.value for cell in row]
        sheet_out.append(data_row)

    # Save the new workbook to a BytesIO object
    output = io.BytesIO()
    workbook_out.save(output)
    output.seek(0)

    response = send_file(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response.headers['Content-Disposition'] = 'attachment; filename=exported_data.xlsx'
    
    return response
    
if __name__ == '__main__':
    app.run(debug=True)

