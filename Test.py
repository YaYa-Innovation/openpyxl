# app.py

from flask import Flask, render_template, request
from openpyxl import load_workbook

app = Flask(__name__)
app.config['UPLOAD_FOLDER'] = 'uploads'

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/display_rows', methods=['POST'])
def display_rows():
    start_row = int(request.form['start_row'])
    end_row = int(request.form['end_row'])

    workbook = load_workbook(f'{app.config["UPLOAD_FOLDER"]}/sample.xlsx')
    sheet = workbook.active

    data = []
    for i in range(start_row, end_row + 1):
        if i >= 1 and i <= sheet.max_row:
            row_data = [sheet.cell(row=i, column=j).value for j in range(1, sheet.max_column + 1)]
            data.append(row_data)

    return render_template('index.html', data=data)

if __name__ == '__main__':
    app.run(debug=True)
<!-- templates/index.html -->

<!DOCTYPE html>
<html lang="en">

<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Display Rows</title>
</head>

<body>

    <h2>Enter Starting and Ending Row Numbers:</h2>

    <form action="/display_rows" method="post">
        Starting Row:<br>
        <input type="number" name="start_row" required><br><br>
        Ending Row:<br>
        <input type="number" name="end_row" required><br><br>
        <input type="submit" value="Submit">
    </form>

    {% if data %}
    <h2>Displaying Rows</h2>
    <table border="1">
        {% for row in data %}
        <tr>
            {% for cell in row %}
            <td>{{ cell }}</td>
            {% endfor %}
        </tr>
        {% endfor %}
    </table>
    {% endif %}

</body>

</html>
# app.py

from flask import Flask, render_template, request, send_file
from openpyxl import Workbook, load_workbook

app = Flask(__name__)
app.config['UPLOAD_FOLDER'] = 'uploads'

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/display_rows', methods=['POST'])
def display_rows():
    start_row = int(request.form['start_row'])
    end_row = int(request.form['end_row'])

    workbook = load_workbook(f'{app.config["UPLOAD_FOLDER"]}/sample.xlsx')
    sheet = workbook.active

    data = []
    for i in range(start_row, end_row + 1):
        if i >= 1 and i <= sheet.max_row:
            row_data = [sheet.cell(row=i, column=j).value for j in range(1, sheet.max_column + 1)]
            data.append(row_data)

    return render_template('index.html', data=data)

@app.route('/export_excel', methods=['POST'])
def export_excel():
    data = request.form.getlist('export_data')
    workbook = Workbook()
    sheet = workbook.active

    for row in data:
        sheet.append(row.split(','))

    file_path = f'{app.config["UPLOAD_FOLDER"]}/exported_data.xlsx'
    workbook.save(file_path)
    return send_file(file_path, as_attachment=True, download_name='exported_data.xlsx')

if __name__ == '__main__':
    app.run(debug=True)
<!-- templates/index.html -->

<!-- ... (existing code remains the same) -->

<form action="/export_excel" method="post">
    <input type="hidden" name="export_data" value="{% for row in data %}{{ row|join(',') }}{% if not loop.last %};{% endif %}{% endfor %}">
    <input type="submit" value="Export to Excel">
</form>
